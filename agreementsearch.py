import os
from pathlib import Path
import openpyxl
from collections import Counter

def find_xlsx_files(root_dir, max_depth=3):
    """
    Find all .xlsx files up to max_depth layers deep.
    """
    xlsx_files = []
    root_path = Path(root_dir)
    
    def search_directory(current_path, current_depth):
        if current_depth > max_depth:
            return
        
        try:
            # Find .xlsx files in current directory
            for item in current_path.iterdir():
                if item.is_file() and item.suffix.lower() == '.xlsx':
                    xlsx_files.append(item)
                elif item.is_dir() and current_depth < max_depth:
                    # Recursively search subdirectories
                    search_directory(item, current_depth + 1)
        except PermissionError:
            pass  # Skip directories we can't access
    
    search_directory(root_path, 0)
    return xlsx_files

def normalize_value(value):
    """
    Convert '_' to '0', return the value as string.
    """
    if value == '_' or value == 0:
        return '0'
    return str(value)

def calculate_cohens_kappa(ratings_data):
    """
    Calculate Cohen's Kappa for two raters.
    Only works when there are exactly 2 raters for all items.
    
    Returns: kappa, observed_agreement, expected_agreement, n_items, categories
    """
    # Filter to only items with exactly 2 raters
    two_rater_data = [ratings for ratings in ratings_data if len(ratings) == 2]
    
    if not two_rater_data:
        return None
    
    n_items = len(two_rater_data)
    
    # Get all unique categories
    all_categories = set()
    for ratings in two_rater_data:
        all_categories.update(ratings)
    categories = sorted(list(all_categories))
    n_categories = len(categories)
    
    # Create confusion matrix
    confusion_matrix = [[0 for _ in categories] for _ in categories]
    
    for ratings in two_rater_data:
        rater1_rating = ratings[0]
        rater2_rating = ratings[1]
        i = categories.index(rater1_rating)
        j = categories.index(rater2_rating)
        confusion_matrix[i][j] += 1
    
    # Calculate observed agreement (Po)
    Po = sum(confusion_matrix[i][i] for i in range(n_categories)) / n_items
    
    # Calculate expected agreement (Pe)
    rater1_marginals = [sum(confusion_matrix[i][j] for j in range(n_categories)) / n_items 
                        for i in range(n_categories)]
    rater2_marginals = [sum(confusion_matrix[i][j] for i in range(n_categories)) / n_items 
                        for j in range(n_categories)]
    
    Pe = sum(rater1_marginals[i] * rater2_marginals[i] for i in range(n_categories))
    
    # Calculate Cohen's Kappa
    if Pe == 1:
        return None
    
    kappa = (Po - Pe) / (1 - Pe)
    
    return kappa, Po, Pe, n_items, categories


def calculate_congers_kappa(ratings_data):
    """
    Calculate Conger's generalized kappa for multiple raters (2 or 3).
    This is a generalization of Cohen's kappa to more than 2 raters.
    
    Returns: kappa, observed_agreement, expected_agreement, n_items, n_raters, categories
    """
    if not ratings_data:
        return None
    
    # Check if we have consistent number of raters (2 or 3)
    rater_counts = set(len(ratings) for ratings in ratings_data)
    if len(rater_counts) > 1:
        print(f"  Warning: Mixed rater counts found: {rater_counts}")
    
    n_items = len(ratings_data)
    
    # Get all unique categories
    all_categories = set()
    for ratings in ratings_data:
        all_categories.update(ratings)
    categories = sorted(list(all_categories))
    n_categories = len(categories)
    
    # For each item, calculate the proportion of rater pairs that agree
    total_pairs_agreeing = 0
    total_pairs = 0
    
    for ratings in ratings_data:
        n_raters = len(ratings)
        # Count agreements among all pairs of raters
        for i in range(n_raters):
            for j in range(i + 1, n_raters):
                total_pairs += 1
                if ratings[i] == ratings[j]:
                    total_pairs_agreeing += 1
    
    # Observed agreement
    Po = total_pairs_agreeing / total_pairs if total_pairs > 0 else 0
    
    # Calculate expected agreement based on marginal probabilities
    # Count how many times each category was assigned across all raters and items
    category_counts = {cat: 0 for cat in categories}
    total_ratings = 0
    
    for ratings in ratings_data:
        for rating in ratings:
            category_counts[rating] += 1
            total_ratings += 1
    
    # Expected agreement is sum of squared proportions
    Pe = sum((count / total_ratings) ** 2 for count in category_counts.values())
    
    # Calculate Conger's Kappa
    if Pe == 1:
        return None
    
    kappa = (Po - Pe) / (1 - Pe)
    
    # Calculate average number of raters
    avg_raters = sum(len(ratings) for ratings in ratings_data) / len(ratings_data)
    
    return kappa, Po, Pe, n_items, avg_raters, categories


def check_all_same(stage_value):
    """
    Check if all sections (separated by colons) are the same after normalization.
    Returns True if all sections are identical, False otherwise.
    """
    if not stage_value or stage_value == '':
        return None  # Skip empty values
    
    parts = str(stage_value).split(':')
    if len(parts) == 0:
        return None
    
    # Normalize all parts
    normalized_parts = [normalize_value(part.strip()) for part in parts]
    
    # Check if all parts are the same
    first_value = normalized_parts[0]
    return all(part == first_value for part in normalized_parts)


def calculate_fleiss_kappa(ratings_data):
    """
    Calculate Fleiss's Kappa for inter-rater agreement.
    
    ratings_data: list of lists, where each inner list contains ratings from multiple raters
                  for a single item (e.g., ['0', '0', '1'] means 3 raters rated one item)
    
    For items with only 2 raters, we assume the third rater agrees with the majority
    (or with the first rater if there's a tie).
    """
    if not ratings_data:
        return None
    
    # Normalize all items to have exactly 3 raters
    normalized_ratings = []
    for ratings in ratings_data:
        if len(ratings) == 3:
            normalized_ratings.append(ratings)
        elif len(ratings) == 2:
            # Assume third rater agrees with the majority, or first rater if tied
            counts = Counter(ratings)
            most_common = counts.most_common(1)[0][0]
            normalized_ratings.append(ratings + [most_common])
        else:
            print(f"Warning: Unexpected number of raters ({len(ratings)}), skipping item")
            continue
    
    if not normalized_ratings:
        return None
    
    n_raters = 3  # Now all items have 3 raters
    n_items = len(normalized_ratings)
    
    # Get all unique categories
    all_categories = set()
    for ratings in normalized_ratings:
        all_categories.update(ratings)
    categories = sorted(list(all_categories))
    n_categories = len(categories)
    
    # Create a matrix: rows = items, columns = categories
    # matrix[i][j] = number of raters who assigned category j to item i
    matrix = []
    for ratings in normalized_ratings:
        counts = Counter(ratings)
        row = [counts.get(cat, 0) for cat in categories]
        matrix.append(row)
    
    # Calculate P_i (proportion of agreement for each item)
    P_values = []
    for row in matrix:
        sum_squares = sum(n_ij ** 2 for n_ij in row)
        P_i = (sum_squares - n_raters) / (n_raters * (n_raters - 1))
        P_values.append(P_i)
    
    # Calculate P_bar (mean proportion of agreement)
    P_bar = sum(P_values) / n_items
    
    # Calculate P_e (expected proportion of agreement by chance)
    p_j_values = []
    for j in range(n_categories):
        sum_for_category = sum(matrix[i][j] for i in range(n_items))
        p_j = sum_for_category / (n_items * n_raters)
        p_j_values.append(p_j)
    
    P_e = sum(p_j ** 2 for p_j in p_j_values)
    
    # Calculate Fleiss's Kappa
    if P_e == 1:
        return None  # All raters always agree, kappa undefined
    
    kappa = (P_bar - P_e) / (1 - P_e)
    
    return kappa, P_bar, P_e, n_items, n_raters, n_categories
    """
    Check if all sections (separated by colons) are the same after normalization.
    Returns True if all sections are identical, False otherwise.
    """
    if not stage_value or stage_value == '':
        return None  # Skip empty values
    
    parts = str(stage_value).split(':')
    if len(parts) == 0:
        return None
    
    # Normalize all parts
    normalized_parts = [normalize_value(part.strip()) for part in parts]
    
    # Check if all parts are the same
    first_value = normalized_parts[0]
    return all(part == first_value for part in normalized_parts)

def analyze_stage_column(file_path):
    """
    Open an xlsx file and analyze the Stage column.
    Returns counts of True and False results, ratings data for Fleiss's Kappa,
    and stage counts.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        results = {'true': 0, 'false': 0, 'skipped': 0}
        ratings_data = []  # Store individual ratings for Fleiss's Kappa
        stage_counts = {'0': 0, '1': 0, '2': 0}  # Count occurrences of each stage
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Find the Stage column
            stage_col_idx = None
            header_row = None
            
            # Search first 10 rows for header
            for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
                if row and any(cell for cell in row):  # Non-empty row
                    for col_idx, cell_value in enumerate(row, start=1):
                        if cell_value and str(cell_value).strip().lower() == 'stage':
                            stage_col_idx = col_idx
                            header_row = row_idx
                            break
                if stage_col_idx:
                    break
            
            if not stage_col_idx:
                print(f"  No 'Stage' column found in sheet '{sheet_name}'")
                continue
            
            print(f"  Found 'Stage' column in sheet '{sheet_name}' at column {stage_col_idx}")
            
            # Iterate through rows ONLY where the row has actual data
            # We'll read the entire row and only process if it's not completely empty
            for row in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                # Check if the row has any non-None values (i.e., it's not a completely empty row)
                if not any(cell is not None for cell in row):
                    continue
                
                # Get the stage value from the appropriate column
                if len(row) >= stage_col_idx:
                    stage_value = row[stage_col_idx - 1]  # Convert to 0-indexed
                    
                    if stage_value is None or str(stage_value).strip() == '':
                        continue  # Skip empty cells
                    
                    # Parse the ratings for Fleiss's Kappa and count stages
                    parts = str(stage_value).split(':')
                    normalized_ratings = [normalize_value(part.strip()) for part in parts]
                    
                    # Count each stage appearance
                    for rating in normalized_ratings:
                        if rating in stage_counts:
                            stage_counts[rating] += 1
                    
                    result = check_all_same(stage_value)
                    
                    if result is None:
                        results['skipped'] += 1
                    else:
                        if result:
                            results['true'] += 1
                        else:
                            results['false'] += 1
                        # Store the ratings for this item
                        ratings_data.append(normalized_ratings)
        
        wb.close()
        return results, ratings_data, stage_counts
    
    except Exception as e:
        print(f"  Error processing file: {e}")
        return {'true': 0, 'false': 0, 'skipped': 0}, [], {'0': 0, '1': 0, '2': 0}

def main():
    root_directory = r"C:\Users\r2d2go\Downloads\agreement"
    
    print(f"Searching for .xlsx files in: {root_directory}")
    print(f"Search depth: up to 3 layers\n")
    
    xlsx_files = find_xlsx_files(root_directory, max_depth=3)
    
    if not xlsx_files:
        print("No .xlsx files found.")
        return
    
    print(f"Found {len(xlsx_files)} .xlsx file(s)\n")
    
    total_true = 0
    total_false = 0
    total_skipped = 0
    all_ratings_data = []
    total_stage_counts = {'0': 0, '1': 0, '2': 0}
    
    for file_path in xlsx_files:
        print(f"Processing: {file_path}")
        results, ratings_data, stage_counts = analyze_stage_column(file_path)
        
        total_true += results['true']
        total_false += results['false']
        total_skipped += results['skipped']
        all_ratings_data.extend(ratings_data)
        
        # Aggregate stage counts
        for stage, count in stage_counts.items():
            total_stage_counts[stage] += count
        
        print(f"  Results - True: {results['true']}, False: {results['false']}, Skipped: {results['skipped']}\n")
    
    print("=" * 60)
    print(f"TOTAL RESULTS:")
    print(f"  All sections same (True): {total_true}")
    print(f"  Sections differ (False): {total_false}")
    print(f"  Skipped/Empty: {total_skipped}")
    print(f"\nSTAGE OCCURRENCES:")
    print(f"  Stage 0: {total_stage_counts['0']}")
    print(f"  Stage 1: {total_stage_counts['1']}")
    print(f"  Stage 2: {total_stage_counts['2']}")
    total_stages = sum(total_stage_counts.values())
    if total_stages > 0:
        print(f"\nSTAGE PERCENTAGES:")
        print(f"  Stage 0: {total_stage_counts['0']/total_stages*100:.2f}%")
        print(f"  Stage 1: {total_stage_counts['1']/total_stages*100:.2f}%")
        print(f"  Stage 2: {total_stage_counts['2']/total_stages*100:.2f}%")
    print("=" * 60)
    
    # Calculate Cohen's Kappa (for 2 raters only)
    if all_ratings_data:
        print("\nCalculating Cohen's Kappa (for items with 2 raters)...")
        cohens_result = calculate_cohens_kappa(all_ratings_data)
        
        if cohens_result:
            kappa, Po, Pe, n_items, categories = cohens_result
            print(f"\nCOHEN'S KAPPA RESULTS:")
            print(f"  Number of items with 2 raters: {n_items}")
            print(f"  Number of categories: {len(categories)}")
            print(f"  Observed agreement (Po): {Po:.4f}")
            print(f"  Expected agreement (Pe): {Pe:.4f}")
            print(f"  Cohen's Kappa (κ): {kappa:.4f}")
            print(f"\nInterpretation:")
            if kappa < 0:
                print("  Poor agreement (worse than chance)")
            elif kappa < 0.20:
                print("  Slight agreement")
            elif kappa < 0.40:
                print("  Fair agreement")
            elif kappa < 0.60:
                print("  Moderate agreement")
            elif kappa < 0.80:
                print("  Substantial agreement")
            else:
                print("  Almost perfect agreement")
        else:
            print("Could not calculate Cohen's Kappa (no items with exactly 2 raters or insufficient data)")
        print("=" * 60)
    
    # Calculate Conger's Kappa (before adding virtual rater)
    if all_ratings_data:
        print("\nCalculating Conger's Generalized Kappa (original data, 2-3 raters)...")
        congers_result = calculate_congers_kappa(all_ratings_data)
        
        if congers_result:
            kappa, Po, Pe, n_items, avg_raters, categories = congers_result
            print(f"\nCONGER'S KAPPA RESULTS:")
            print(f"  Number of items: {n_items}")
            print(f"  Average number of raters: {avg_raters:.2f}")
            print(f"  Number of categories: {len(categories)}")
            print(f"  Observed agreement (Po): {Po:.4f}")
            print(f"  Expected agreement (Pe): {Pe:.4f}")
            print(f"  Conger's Kappa (κ): {kappa:.4f}")
            print(f"\nInterpretation:")
            if kappa < 0:
                print("  Poor agreement (worse than chance)")
            elif kappa < 0.20:
                print("  Slight agreement")
            elif kappa < 0.40:
                print("  Fair agreement")
            elif kappa < 0.60:
                print("  Moderate agreement")
            elif kappa < 0.80:
                print("  Substantial agreement")
            else:
                print("  Almost perfect agreement")
        else:
            print("Could not calculate Conger's Kappa (insufficient or invalid data)")
        print("=" * 60)
    
    # Calculate Fleiss's Kappa
    if all_ratings_data:
        print("\nCalculating Fleiss's Kappa (normalized to 3 raters with virtual rater)...")
        kappa_result = calculate_fleiss_kappa(all_ratings_data)
        
        if kappa_result:
            kappa, P_bar, P_e, n_items, n_raters, n_categories = kappa_result
            print(f"\nFLEISS'S KAPPA RESULTS:")
            print(f"  Number of items: {n_items}")
            print(f"  Number of raters: {n_raters}")
            print(f"  Number of categories: {n_categories}")
            print(f"  Observed agreement (P̄): {P_bar:.4f}")
            print(f"  Expected agreement (Pₑ): {P_e:.4f}")
            print(f"  Fleiss's Kappa (κ): {kappa:.4f}")
            print(f"\nInterpretation:")
            if kappa < 0:
                print("  Poor agreement (worse than chance)")
            elif kappa < 0.20:
                print("  Slight agreement")
            elif kappa < 0.40:
                print("  Fair agreement")
            elif kappa < 0.60:
                print("  Moderate agreement")
            elif kappa < 0.80:
                print("  Substantial agreement")
            else:
                print("  Almost perfect agreement")
        else:
            print("Could not calculate Fleiss's Kappa (insufficient or invalid data)")
    else:
        print("\nNo valid ratings data found for Fleiss's Kappa calculation")
    print("=" * 60)

if __name__ == "__main__":
    main()